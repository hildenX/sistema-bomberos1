"""
Views para Carga Masiva de Voluntarios desde Excel
"""
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import transaction
from .models import Voluntario
from .utils import VoluntarioUtils
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import time
from datetime import datetime
from .permissions import autorizar_request, RolBomberos


ROLES_CARGA_MASIVA = (RolBomberos.SUPER_ADMIN,)


def _autorizar_admin(request):
    autorizado, response = autorizar_request(request, roles=ROLES_CARGA_MASIVA)
    return None if autorizado else response


@require_http_methods(["GET"])
def descargar_plantilla_masiva(request):
    """
    Genera y descarga plantilla Excel para carga masiva de voluntarios
    GET /api/voluntarios/descargar-plantilla-masiva/
    """
    auth_response = _autorizar_admin(request)
    if auth_response:
        return auth_response

    print("[CARGA MASIVA] Generando plantilla Excel...")
    
    # Crear workbook
    wb = Workbook()
    
    # ==================== UNA SOLA HOJA: VOLUNTARIOS ====================
    ws = wb.active
    ws.title = "VOLUNTARIOS"
    
    # Headers - TODOS LOS CAMPOS DEL FORMULARIO DE AGREGAR BOMBERO
    headers = [
        'CLAVE', 'PRIMER_NOMBRE', 'SEGUNDO_NOMBRE', 'TERCER_NOMBRE',
        'PRIMER_APELLIDO', 'SEGUNDO_APELLIDO', 'RUT', 'FECHA_NAC',
        'TELEFONO', 'EMAIL', 'DOMICILIO', 'PROFESION', 'GRUPO_SANG',
        'COMPANIA', 'FECHA_INGRESO', 'NRO_REGISTRO',
        'NOMBRE_PRIMER_PADRINO', 'NOMBRE_SEGUNDO_PADRINO',
        'OTROS_CUERPOS', 'COMPANIA_OPCIONAL', 'DESDE', 'HASTA',
        'ESTADO'
    ]
    
    # Estilo para headers
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ejemplos (3 filas con TODOS los campos - 23 columnas) - RUTs VÁLIDOS
    ejemplos = [
        ['B001', 'Cristian', 'Alejandro', 'José', 'Vera', 'Arriagada', '12.345.678-5', '1990-05-15',
         '+56912345678', 'cristian@mail.cl', 'Calle 1 #123', 'Ingeniero', 'O+', 'Quinta', 
         '2015-01-10', 'REG-001', 'Juan Pérez González', 'María López Silva', 
         'Valparaíso', 'Segunda', '2010', '2015', 'activo'],
        ['B002', 'María', 'Fernanda', '', 'González', 'Rojas', '18.765.432-8', '1985-08-20',
         '+56987654321', 'maria@mail.cl', 'Av. 2 #456', 'Enfermera', 'A+', 'Primera',
         '2010-06-15', 'REG-002', 'Pedro Silva Torres', 'Ana Muñoz Castro', 
         '', '', '', '', 'activo'],
        ['B003', 'Carlos', '', '', 'Ramírez', 'Silva', '16.223.344-6', '1995-12-10',
         '+56911223344', 'carlos@mail.cl', 'Pasaje 3 #789', 'Profesor', 'B+', 'Tercera',
         '2020-03-20', 'REG-003', 'Luis Díaz Rojas', 'Carmen Fuentes Pérez',
         'Concepción', 'Primera', '2015', '2020', 'fallecido'],
    ]
    
    # Escribir ejemplos
    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col_num, value in enumerate(ejemplo, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Ajustar anchos de columna (23 columnas totales)
    column_widths = [12, 15, 15, 15, 15, 15, 15, 13, 15, 20, 25, 15, 12, 12, 15, 15, 25, 25, 18, 18, 10, 10, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ==================== HOJA DE INSTRUCCIONES ====================
    ws_instruc = wb.create_sheet("INSTRUCCIONES")
    ws_instruc.column_dimensions['A'].width = 80
    
    instrucciones = [
        ("📋 INSTRUCCIONES DE USO - CARGA MASIVA DE VOLUNTARIOS", True),
        ("", False),
        ("✅ CAMPOS OBLIGATORIOS:", True),
        ("   • CLAVE (único, ej: B001)", False),
        ("   • PRIMER_NOMBRE, PRIMER_APELLIDO, SEGUNDO_APELLIDO", False),
        ("   • RUT (formato: 12.345.678-9 con guión)", False),
        ("   • FECHA_INGRESO (formato: YYYY-MM-DD, ej: 2020-01-15)", False),
        ("   • NOMBRE_PRIMER_PADRINO, NOMBRE_SEGUNDO_PADRINO", False),
        ("", False),
        ("⚠️ ESTADO (solo estos 3 valores):", True),
        ("   • activo", False),
        ("   • fallecido", False),
        ("   • martir", False),
        ("", False),
        ("📝 FORMATOS:", True),
        ("   • Fechas: YYYY-MM-DD (2024-01-15)", False),
        ("   • RUT: 12.345.678-9 (con puntos y guión)", False),
        ("   • Teléfono: +56912345678", False),
        ("   • Compañía: Primera, Segunda, Tercera, Cuarta, Quinta", False),
        ("   • Grupo Sanguíneo: A+, A-, B+, B-, AB+, AB-, O+, O-", False),
        ("", False),
        ("📌 CAMPOS OPCIONALES:", True),
        ("   • SEGUNDO_NOMBRE, TERCER_NOMBRE", False),
        ("   • FECHA_NAC, TELEFONO, EMAIL, DOMICILIO", False),
        ("   • PROFESION, GRUPO_SANG, NRO_REGISTRO", False),
        ("   • OTROS_CUERPOS, COMPANIA_OPCIONAL, DESDE, HASTA", False),
        ("", False),
        ("📊 LÍMITES:", True),
        ("   • Máximo 500 voluntarios por archivo", False),
        ("", False),
        ("💡 PASOS:", True),
        ("   1. Opción A: Usa los ejemplos directamente (filas 2, 3, 4)", False),
        ("   2. Opción B: Elimina ejemplos y llena tus propios datos", False),
        ("   3. Guardar el archivo", False),
        ("   4. Subir en el panel de Carga Masiva", False),
    ]
    
    for row_num, (texto, es_titulo) in enumerate(instrucciones, 1):
        cell = ws_instruc.cell(row=row_num, column=1, value=texto)
        if es_titulo:
            cell.font = Font(bold=True, size=12, color="C8102E")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Generar archivo Excel en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Preparar respuesta
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Plantilla_Bomberos.xlsx"'
    
    print("[CARGA MASIVA] ✅ Plantilla generada exitosamente")
    return response


@csrf_exempt
@require_http_methods(["POST"])
def importar_masiva(request):
    """
    Importa voluntarios desde archivo Excel
    POST /api/voluntarios/importar-masiva/
    Body: FormData con archivo Excel
    """
    auth_response = _autorizar_admin(request)
    if auth_response:
        return auth_response

    print("[CARGA MASIVA] Iniciando importación...")
    
    inicio = time.time()
    
    # Validar archivo
    if 'archivo' not in request.FILES:
        return JsonResponse({'error': 'No se recibió ningún archivo'}, status=400)
    
    archivo = request.FILES['archivo']
    
    # Validar extensión
    if not (archivo.name.endswith('.xlsx') or archivo.name.endswith('.xls')):
        return JsonResponse({'error': 'El archivo debe ser .xlsx o .xls'}, status=400)
    
    try:
        # Cargar workbook
        wb = load_workbook(archivo, data_only=True)
        
        # Leer hoja principal (VOLUNTARIOS o la activa)
        ws = wb['VOLUNTARIOS'] if 'VOLUNTARIOS' in wb.sheetnames else wb.active
        
        # Procesar filas
        voluntarios_creados = []
        errores = []
        total_procesados = 0
        
        # Saltar solo el header (fila 1), leer desde fila 2
        for row_num in range(2, ws.max_row + 1):
            try:
                # Leer TODOS los datos de UNA SOLA FILA (23 columnas)
                clave = ws.cell(row=row_num, column=1).value
                
                # Si la clave está vacía, terminar
                if not clave:
                    continue
                
                total_procesados += 1
                
                # Validar clave única
                if Voluntario.objects.filter(clave_bombero=clave).exists():
                    errores.append({
                        'fila': row_num,
                        'error': f'La clave {clave} ya existe en el sistema'
                    })
                    continue
                
                # Leer nombres (columnas 2-4)
                primer_nombre = ws.cell(row=row_num, column=2).value
                segundo_nombre = ws.cell(row=row_num, column=3).value
                tercer_nombre = ws.cell(row=row_num, column=4).value
                
                # Leer apellidos (columnas 5-6)
                primer_apellido = ws.cell(row=row_num, column=5).value
                segundo_apellido = ws.cell(row=row_num, column=6).value
                
                # Leer datos personales (columnas 7-16)
                rut = ws.cell(row=row_num, column=7).value
                fecha_nac = ws.cell(row=row_num, column=8).value
                telefono = ws.cell(row=row_num, column=9).value
                email = ws.cell(row=row_num, column=10).value
                domicilio = ws.cell(row=row_num, column=11).value
                profesion = ws.cell(row=row_num, column=12).value
                grupo_sang = ws.cell(row=row_num, column=13).value
                compania = ws.cell(row=row_num, column=14).value
                fecha_ingreso = ws.cell(row=row_num, column=15).value
                nro_registro = ws.cell(row=row_num, column=16).value
                
                # Leer padrinos (columnas 17-18)
                primer_padrino = ws.cell(row=row_num, column=17).value
                segundo_padrino = ws.cell(row=row_num, column=18).value
                
                # Leer otros cuerpos (columnas 19-22)
                otros_cuerpos = ws.cell(row=row_num, column=19).value
                comp_opcion = ws.cell(row=row_num, column=20).value
                desde = ws.cell(row=row_num, column=21).value
                hasta = ws.cell(row=row_num, column=22).value
                
                # Leer estado (columna 23) - SOLO 3 VALORES PERMITIDOS
                estado_raw = ws.cell(row=row_num, column=23).value
                estado = str(estado_raw).lower().strip() if estado_raw else 'activo'
                
                # Validar que solo sean los 3 estados permitidos
                if estado not in ['activo', 'fallecido', 'martir']:
                    errores.append({
                        'fila': row_num,
                        'error': f'Estado inválido: {estado_raw}. Solo se permiten: activo, fallecido, martir'
                    })
                    continue
                
                # Validaciones obligatorias
                if not all([primer_nombre, primer_apellido, segundo_apellido, rut, fecha_ingreso, primer_padrino, segundo_padrino]):
                    errores.append({
                        'fila': row_num,
                        'error': 'Faltan campos obligatorios (Primer Nombre, Primer Apellido, Segundo Apellido, RUT, Fecha Ingreso, Padrinos)'
                    })
                    continue
                
                # Concatenar nombres completos
                nombre_completo = f"{primer_nombre or ''} {segundo_nombre or ''} {tercer_nombre or ''}".strip()
                
                # Validar RUT
                rut_limpio = str(rut).replace('.', '').replace('-', '').strip()
                if not VoluntarioUtils.validar_rut(rut_limpio):
                    errores.append({
                        'fila': row_num,
                        'error': f'RUT inválido: {rut}'
                    })
                    continue
                
                # Formatear RUT con puntos y guión (12.345.678-9)
                rut_formateado = VoluntarioUtils.formatear_rut(rut_limpio)
                
                # Verificar RUT único
                if Voluntario.objects.filter(rut=rut_formateado).exists():
                    errores.append({
                        'fila': row_num,
                        'error': f'El RUT {rut} ya existe en el sistema'
                    })
                    continue
                
                # Convertir fechas
                if isinstance(fecha_ingreso, str):
                    fecha_ingreso = datetime.strptime(fecha_ingreso, '%Y-%m-%d').date()
                
                if fecha_nac and isinstance(fecha_nac, str):
                    fecha_nac = datetime.strptime(fecha_nac, '%Y-%m-%d').date()
                
                # Los padrinos y estado ya fueron leídos arriba de la misma fila
                
                # Crear voluntario
                with transaction.atomic():
                    voluntario = Voluntario.objects.create(
                        clave_bombero=clave,
                        nombre=nombre_completo,
                        apellido_paterno=primer_apellido,
                        apellido_materno=segundo_apellido,
                        rut=rut_formateado,
                        fecha_nacimiento=fecha_nac,
                        telefono=telefono or '',
                        email=email or '',
                        domicilio=domicilio or '',
                        profesion=profesion or '',
                        grupo_sanguineo=grupo_sang or '',
                        compania=compania or 'Quinta',
                        fecha_ingreso=fecha_ingreso,
                        nro_registro=nro_registro or '',
                        nombre_primer_padrino=primer_padrino,
                        nombre_segundo_padrino=segundo_padrino,
                        otros_cuerpos=otros_cuerpos or '',
                        compania_opcional=comp_opcion or '',
                        desde=desde or '',
                        hasta=hasta or '',
                        estado_bombero=estado
                    )
                    voluntarios_creados.append(voluntario.id)
                
                print(f"[CARGA MASIVA] ✅ Fila {row_num}: {clave} - {nombre_completo} {primer_apellido}")
                
            except Exception as e:
                print(f"[CARGA MASIVA] ❌ Error fila {row_num}: {str(e)}")
                errores.append({
                    'fila': row_num,
                    'error': str(e)
                })
        
        # Calcular tiempo
        tiempo = f"{time.time() - inicio:.2f} seg"
        
        # Respuesta
        return JsonResponse({
            'success': True,
            'mensaje': f'{len(voluntarios_creados)} voluntarios importados correctamente',
            'total_procesados': total_procesados,
            'exitosos': len(voluntarios_creados),
            'errores': errores,
            'tiempo': tiempo
        })
        
    except Exception as e:
        print(f"[CARGA MASIVA] ❌ Error general: {str(e)}")
        return JsonResponse({
            'error': f'Error al procesar archivo: {str(e)}'
        }, status=500)
