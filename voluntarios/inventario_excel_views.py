"""
Vista para exportar el Inventario del Pañol a Excel
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

from .models import ItemInventario
from .permissions import requiere_permiso


@requiere_permiso('inventario', 'view')
def exportar_inventario_excel(request):
    """
    Genera y descarga el inventario completo del pañol como Excel.
    GET /api/voluntarios/inventario/exportar-excel/
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTARIO PAÑOL"

    headers = [
        'NOMBRE', 'CATEGORIA', 'CANTIDAD', 'UNIDAD', 'MARCA', 'TAMAÑO',
        'ESTADO', 'N° DE SERIE', 'RESPONSABLE', 'UBICACION', 'OBSERVACIONES'
    ]

    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    items = ItemInventario.objects.all().order_by('categoria', 'nombre')
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.nombre)
        ws.cell(row=row_num, column=2, value=item.categoria)
        ws.cell(row=row_num, column=3, value=float(item.cantidad))
        ws.cell(row=row_num, column=4, value=item.unidad)
        ws.cell(row=row_num, column=5, value=item.marca)
        ws.cell(row=row_num, column=6, value=item.tamano)
        ws.cell(row=row_num, column=7, value=item.get_estado_display() if item.estado else '')
        ws.cell(row=row_num, column=8, value=item.numero_serie)
        ws.cell(row=row_num, column=9, value=item.responsable)
        ws.cell(row=row_num, column=10, value=item.ubicacion)
        ws.cell(row=row_num, column=11, value=item.observaciones)

    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Inventario_Panol.xlsx"'
    return response
